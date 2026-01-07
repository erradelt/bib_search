from docx import Document
from PyPDF2 import PdfReader
from openpyxl import load_workbook

class EXLhandler:
    def __init__(self, doc_path, searchterm):
        self.doc_path = doc_path
        self.searchterm = searchterm
        self.sheetnames = []
        self.rows = {}
        self.found_items = {}

    def read_doc(self):
        return(load_workbook(self.doc_path))

    def exl_to_sheets(self):
        for sheet in self.read_doc().sheetnames:
            self.sheetnames.append(sheet)
        return self.sheetnames

    def sheets_to_rows(self):
        for sheet in self.exl_to_sheets():
            temp = []
            self.ws = self.read_doc()[sheet]
            for row in self.ws.iter_rows(min_row=2, values_only=True):
                temp.append(row)
            self.rows[sheet]=temp
        return self.rows

    def searchterm_finder(self):
        for key, value in self.sheets_to_rows().items():
            for list in value:
                for item in list:
                    if self.searchterm == item:
                        self.found_items[f'sheet: {key}'] = item
        return self.found_items

class DOChandler:
    def __init__(self, doc_path, searchterm):
        self.doc_path = doc_path
        self.searchterm = searchterm

    def read_doc(self):
        return(Document(self.doc_path))

    def doc_dissector(self):
        self.para_dict = {}
        i=1
        for para in self.read_doc().paragraphs:
            self.para_dict[f"absatz {i}"]=para.text
            i += 1
        return self.para_dict

    def searchterm_finder(self):
        #key = page of document
        #value = text of corresponding page
        self.found_dict = {}
        for key, value in self.doc_dissector().items():
            found_word = value.find(self.searchterm)
            if found_word != - 1: # -1 means searchterm not found in value
                start = max(0, found_word-75)
                end = min(len(value), found_word+len(self.searchterm)+75)
                self.found_dict[key]=f'...{value[start:end]}...'.replace('\n','')
        return self.found_dict

class PDFhandler:
    def __init__(self, doc_path, searchterm):
        self.doc_path = doc_path
        self.searchterm = searchterm

    def read_doc(self):
        return(PdfReader(self.doc_path))

    def doc_dissector(self):
        self.page_dict = {}
        i=1
        for page in self.read_doc().pages:
            self.page_dict[f"seite {i}"]=page.extract_text()
            i += 1
        return self.page_dict

    def searchterm_finder(self):
        #key = page of document
        #value = text of corresponding page
        self.found_dict = {}
        for key, value in self.doc_dissector().items():
            found_word = value.find(self.searchterm)
            if found_word != - 1: # -1 means searchterm not found in value
                start = max(0, found_word-75)
                end = min(len(value), found_word+len(self.searchterm)+75)
                self.found_dict[key]=f'...{value[start:end]}...'.replace('\n','')
        return self.found_dict

exltest = EXLhandler('/home/robert/Downloads/beispiel_mehrere_tabellenblaetter.xlsx','Wartung')
print(exltest.searchterm_finder())
